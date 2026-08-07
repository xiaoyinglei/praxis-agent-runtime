from __future__ import annotations

import asyncio
import csv
import hashlib
import json
import math
from collections.abc import Mapping, Sequence
from datetime import date, datetime, time
from pathlib import Path
from typing import Literal

from pydantic import BaseModel, ConfigDict, Field

from agent_runtime.tools.tool import (
    CancellationMode,
    InterruptBehavior,
    JsonValue,
    NormalizedToolOutput,
    ResolvedToolUse,
    Tool,
    ToolDefinition,
    ToolEffect,
    ToolTarget,
    json_schema_output,
    pydantic_input,
)
from agent_runtime.workspace import WorkspaceRuntime

_MAX_INSPECTABLE_BYTES = 200 * 1024 * 1024
_MAX_CELL_CHARACTERS = 2_000


class InspectDataFileInput(BaseModel):
    model_config = ConfigDict(extra="forbid")

    path: str = Field(
        min_length=1,
        max_length=4096,
        description=(
            "Workspace-relative .xlsx, .xlsm, .pdf, .csv, .tsv, or .json "
            "file to inspect."
        ),
    )
    max_rows: int = Field(
        default=10,
        ge=1,
        le=100,
        description="Maximum data rows returned for each table or worksheet.",
    )
    max_sheets: int = Field(
        default=20,
        ge=1,
        le=100,
        description="Maximum worksheets returned from one workbook.",
    )
    max_pages: int = Field(
        default=10,
        ge=1,
        le=100,
        description="Maximum PDF pages whose extracted text is returned.",
    )
    max_characters: int = Field(
        default=8_000,
        ge=100,
        le=100_000,
        description=(
            "Maximum characters returned per PDF page or JSON preview."
        ),
    )


class TablePreview(BaseModel):
    model_config = ConfigDict(extra="forbid")

    name: str
    headers: list[str]
    rows: list[list[str]]
    row_count: int = Field(ge=0)
    column_count: int = Field(ge=0)
    truncated: bool


class PagePreview(BaseModel):
    model_config = ConfigDict(extra="forbid")

    page_number: int = Field(ge=1)
    text: str
    truncated: bool


class InspectDataFileOutput(BaseModel):
    model_config = ConfigDict(extra="forbid")

    path: str
    format: Literal["xlsx", "xlsm", "pdf", "csv", "tsv", "json", "unsupported"]
    size_bytes: int = Field(ge=0)
    sha256: str
    valid: bool
    summary: str
    tables: list[TablePreview]
    sheet_names: list[str]
    page_count: int | None
    pages: list[PagePreview]
    json_type: str | None
    json_keys: list[str]
    json_preview: str | None
    truncated: bool
    error: str | None


_INSPECT_INPUT_SCHEMA, _validate_inspect_input = pydantic_input(
    InspectDataFileInput
)
_INSPECT_OUTPUT_SCHEMA, _unused_inspect_output_validator = pydantic_input(
    InspectDataFileOutput
)


def create_inspect_data_file_tool(workspace: WorkspaceRuntime) -> Tool:
    async def run(
        arguments: Mapping[str, JsonValue],
    ) -> InspectDataFileOutput:
        request = InspectDataFileInput.model_validate(arguments)
        return await asyncio.to_thread(_inspect_data_file, workspace, request)

    return Tool(
        definition=ToolDefinition(
            name="inspect_data_file",
            description=(
                "Inspect a workspace spreadsheet, PDF, CSV, TSV, or JSON file "
                "without returning raw binary bytes. The result includes a "
                "runtime-computed SHA-256 plus bounded worksheet rows, extracted "
                "PDF text, delimited-table rows, or a JSON preview. Use this before "
                "analysis and once after execute_python creates a data artifact. A "
                "successful inspection of the exact generated path verifies that "
                "artifact's file structure and content; then finish instead of "
                "reading the original binary again or repeating the inspection."
            ),
            input_schema=_INSPECT_INPUT_SCHEMA,
        ),
        validate_input=_validate_inspect_input,
        run=run,
        normalize_output=_normalize_inspect_output,
        output_schema=_INSPECT_OUTPUT_SCHEMA,
        static_effects=frozenset({ToolEffect.READ_WORKSPACE}),
        resolve_use=lambda arguments: _resolve_inspect_use(workspace, arguments),
        execution_revision="builtin-inspect-data-file-v1",
        idempotent=True,
        concurrency_safe=True,
        cancellation_mode=CancellationMode.COOPERATIVE,
        interrupt_behavior=InterruptBehavior.CANCEL,
        timeout_seconds=60.0,
        max_model_output_bytes=500_000,
    )


def _inspect_data_file(
    workspace: WorkspaceRuntime,
    request: InspectDataFileInput,
) -> InspectDataFileOutput:
    target = workspace.ensure_within_workspace(
        workspace.resolve_path(request.path)
    )
    if not target.is_file():
        raise FileNotFoundError(
            f"workspace data file not found: {request.path}"
        )
    size_bytes = target.stat().st_size
    if size_bytes > _MAX_INSPECTABLE_BYTES:
        return _inspection_failure(
            request=request,
            target=target,
            file_format=_format_from_path(target),
            size_bytes=size_bytes,
            error=(
                "data file exceeds the 200 MiB inspection limit; create a "
                "bounded sample with execute_python"
            ),
        )

    file_format = _format_from_path(target)
    try:
        if file_format == "xlsx":
            return _inspect_workbook(
                target,
                request=request,
                file_format="xlsx",
                size_bytes=size_bytes,
            )
        if file_format == "xlsm":
            return _inspect_workbook(
                target,
                request=request,
                file_format="xlsm",
                size_bytes=size_bytes,
            )
        if file_format == "csv":
            return _inspect_delimited(
                target,
                request=request,
                file_format="csv",
                size_bytes=size_bytes,
            )
        if file_format == "tsv":
            return _inspect_delimited(
                target,
                request=request,
                file_format="tsv",
                size_bytes=size_bytes,
            )
        if file_format == "pdf":
            return _inspect_pdf(
                target,
                request=request,
                size_bytes=size_bytes,
            )
        if file_format == "json":
            return _inspect_json(
                target,
                request=request,
                size_bytes=size_bytes,
            )
        return _inspection_failure(
            request=request,
            target=target,
            file_format="unsupported",
            size_bytes=size_bytes,
            error=(
                "unsupported data format; expected .xlsx, .xlsm, .pdf, .csv, "
                ".tsv, or .json"
            ),
        )
    except Exception as exc:
        return _inspection_failure(
            request=request,
            target=target,
            file_format=file_format,
            size_bytes=size_bytes,
            error=_bounded_error(exc),
        )


def _inspect_workbook(
    target: Path,
    *,
    request: InspectDataFileInput,
    file_format: Literal["xlsx", "xlsm"],
    size_bytes: int,
) -> InspectDataFileOutput:
    from openpyxl import load_workbook

    workbook = load_workbook(
        target,
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    try:
        sheet_names = list(workbook.sheetnames)
        previews: list[TablePreview] = []
        for worksheet in workbook.worksheets[: request.max_sheets]:
            row_iterator = worksheet.iter_rows(values_only=True)
            first_row = next(row_iterator, ())
            column_count = max(worksheet.max_column or 0, len(first_row))
            headers = [
                _cell_text(value) or f"column_{index}"
                for index, value in enumerate(first_row, start=1)
            ]
            if len(headers) < column_count:
                headers.extend(
                    f"column_{index}"
                    for index in range(len(headers) + 1, column_count + 1)
                )
            rows: list[list[str]] = []
            for values in row_iterator:
                if len(rows) >= request.max_rows:
                    break
                normalized = [_cell_text(value) for value in values]
                if len(normalized) < column_count:
                    normalized.extend("" for _ in range(column_count - len(normalized)))
                rows.append(normalized[:column_count])
            row_count = max((worksheet.max_row or 0) - (1 if first_row else 0), 0)
            previews.append(
                TablePreview(
                    name=worksheet.title,
                    headers=headers[:column_count],
                    rows=rows,
                    row_count=row_count,
                    column_count=column_count,
                    truncated=row_count > len(rows),
                )
            )
        truncated = len(sheet_names) > len(previews) or any(
            preview.truncated for preview in previews
        )
        return InspectDataFileOutput(
            path=request.path,
            format=file_format,
            size_bytes=size_bytes,
            sha256=_file_sha256(target),
            valid=True,
            summary=(
                f"Valid {file_format.upper()} workbook with "
                f"{len(sheet_names)} worksheet(s)."
            ),
            tables=previews,
            sheet_names=sheet_names,
            page_count=None,
            pages=[],
            json_type=None,
            json_keys=[],
            json_preview=None,
            truncated=truncated,
            error=None,
        )
    finally:
        workbook.close()


def _inspect_delimited(
    target: Path,
    *,
    request: InspectDataFileInput,
    file_format: Literal["csv", "tsv"],
    size_bytes: int,
) -> InspectDataFileOutput:
    delimiter = "\t" if file_format == "tsv" else ","
    with target.open("r", encoding="utf-8-sig", errors="replace", newline="") as stream:
        sample = stream.read(16_384)
        stream.seek(0)
        if file_format == "csv" and sample:
            try:
                delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
            except csv.Error:
                pass
        reader = csv.reader(stream, delimiter=delimiter)
        header_row = next(reader, [])
        headers = [
            _cell_text(value) or f"column_{index}"
            for index, value in enumerate(header_row, start=1)
        ]
        rows: list[list[str]] = []
        row_count = 0
        column_count = len(headers)
        for values in reader:
            row_count += 1
            column_count = max(column_count, len(values))
            if len(rows) < request.max_rows:
                rows.append([_cell_text(value) for value in values])

    if len(headers) < column_count:
        headers.extend(
            f"column_{index}"
            for index in range(len(headers) + 1, column_count + 1)
        )
    for row in rows:
        row.extend("" for _ in range(column_count - len(row)))
    table = TablePreview(
        name=target.name,
        headers=headers,
        rows=rows,
        row_count=row_count,
        column_count=column_count,
        truncated=row_count > len(rows),
    )
    return InspectDataFileOutput(
        path=request.path,
        format=file_format,
        size_bytes=size_bytes,
        sha256=_file_sha256(target),
        valid=True,
        summary=(
            f"Valid {file_format.upper()} table with {row_count} data row(s) "
            f"and {column_count} column(s)."
        ),
        tables=[table],
        sheet_names=[],
        page_count=None,
        pages=[],
        json_type=None,
        json_keys=[],
        json_preview=None,
        truncated=table.truncated,
        error=None,
    )


def _inspect_pdf(
    target: Path,
    *,
    request: InspectDataFileInput,
    size_bytes: int,
) -> InspectDataFileOutput:
    import fitz  # type: ignore[import-untyped]

    document = fitz.open(target)
    try:
        if document.needs_pass:
            raise ValueError("encrypted PDF requires a password")
        page_count = document.page_count
        pages: list[PagePreview] = []
        for page_index in range(min(page_count, request.max_pages)):
            text = document.load_page(page_index).get_text("text")
            truncated = len(text) > request.max_characters
            pages.append(
                PagePreview(
                    page_number=page_index + 1,
                    text=text[: request.max_characters],
                    truncated=truncated,
                )
            )
        truncated = page_count > len(pages) or any(page.truncated for page in pages)
        return InspectDataFileOutput(
            path=request.path,
            format="pdf",
            size_bytes=size_bytes,
            sha256=_file_sha256(target),
            valid=True,
            summary=f"Valid PDF with {page_count} page(s).",
            tables=[],
            sheet_names=[],
            page_count=page_count,
            pages=pages,
            json_type=None,
            json_keys=[],
            json_preview=None,
            truncated=truncated,
            error=None,
        )
    finally:
        document.close()


def _inspect_json(
    target: Path,
    *,
    request: InspectDataFileInput,
    size_bytes: int,
) -> InspectDataFileOutput:
    with target.open("r", encoding="utf-8-sig") as stream:
        value = json.load(stream)
    json_type = _json_type(value)
    json_keys = sorted(str(key) for key in value) if isinstance(value, dict) else []
    rendered = json.dumps(
        value,
        ensure_ascii=False,
        indent=2,
        allow_nan=False,
        default=_json_default,
    )
    truncated = len(rendered) > request.max_characters
    return InspectDataFileOutput(
        path=request.path,
        format="json",
        size_bytes=size_bytes,
        sha256=_file_sha256(target),
        valid=True,
        summary=(
            f"Valid JSON {json_type}"
            + (f" with {len(value)} item(s)." if isinstance(value, Sequence) and not isinstance(value, str) else ".")
        ),
        tables=[],
        sheet_names=[],
        page_count=None,
        pages=[],
        json_type=json_type,
        json_keys=json_keys,
        json_preview=rendered[: request.max_characters],
        truncated=truncated,
        error=None,
    )


def _inspection_failure(
    *,
    request: InspectDataFileInput,
    target: Path,
    file_format: Literal[
        "xlsx", "xlsm", "pdf", "csv", "tsv", "json", "unsupported"
    ],
    size_bytes: int,
    error: str,
) -> InspectDataFileOutput:
    return InspectDataFileOutput(
        path=request.path,
        format=file_format,
        size_bytes=size_bytes,
        sha256=_file_sha256(target),
        valid=False,
        summary="Data file inspection failed.",
        tables=[],
        sheet_names=[],
        page_count=None,
        pages=[],
        json_type=None,
        json_keys=[],
        json_preview=None,
        truncated=False,
        error=error,
    )


def _normalize_inspect_output(raw: object) -> NormalizedToolOutput:
    validated = InspectDataFileOutput.model_validate(raw)
    structured = json_schema_output(
        _INSPECT_OUTPUT_SCHEMA,
        validated.model_dump(mode="json"),
    )
    if not validated.valid:
        return NormalizedToolOutput(
            structured_content=structured,
            is_error=True,
            error_code="data_inspection_failed",
            error_message=validated.error or "data file inspection failed",
            retryable=False,
        )
    return NormalizedToolOutput(structured_content=structured)


def _resolve_inspect_use(
    workspace: WorkspaceRuntime,
    arguments: Mapping[str, JsonValue],
) -> ResolvedToolUse:
    target = workspace.ensure_within_workspace(
        workspace.resolve_path(str(arguments["path"]))
    )
    return ResolvedToolUse(
        effects=frozenset({ToolEffect.READ_WORKSPACE}),
        targets=(ToolTarget(kind="workspace_path", value=str(target)),),
    )


def _format_from_path(
    path: Path,
) -> Literal["xlsx", "xlsm", "pdf", "csv", "tsv", "json", "unsupported"]:
    suffix = path.suffix.casefold()
    if suffix == ".xlsx":
        return "xlsx"
    if suffix == ".xlsm":
        return "xlsm"
    if suffix == ".pdf":
        return "pdf"
    if suffix == ".csv":
        return "csv"
    if suffix == ".tsv":
        return "tsv"
    if suffix == ".json":
        return "json"
    return "unsupported"


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date, time)):
        rendered = value.isoformat()
    elif isinstance(value, float) and not math.isfinite(value):
        rendered = str(value)
    elif isinstance(value, bytes):
        rendered = f"<binary value: {len(value)} bytes>"
    else:
        rendered = str(value)
    if len(rendered) <= _MAX_CELL_CHARACTERS:
        return rendered
    return rendered[: _MAX_CELL_CHARACTERS - 1] + "…"


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        while chunk := stream.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def _bounded_error(error: Exception) -> str:
    message = " ".join(str(error).split()) or type(error).__name__
    return message[:1_000]


def _json_type(value: object) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, dict):
        return "object"
    if isinstance(value, list):
        return "array"
    if isinstance(value, str):
        return "string"
    if isinstance(value, (int, float)):
        return "number"
    return type(value).__name__


def _json_default(value: object) -> str:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


__all__ = [
    "InspectDataFileInput",
    "InspectDataFileOutput",
    "PagePreview",
    "TablePreview",
    "create_inspect_data_file_tool",
]
